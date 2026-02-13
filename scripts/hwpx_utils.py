#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
hwpx_utils.py — HWPX 서식 자동 채우기 공용 유틸리티
====================================================
이 모듈은 한글(HWPX) 서식에 엑셀 데이터를 채워넣는 작업에 필요한
공통 함수들을 모아둔 라이브러리입니다.

주요 기능:
  - HWPX 파일 읽기/쓰기 (ZIP + XML 조작)
  - cellAddr 좌표 기반 셀 채우기
  - 엑셀 데이터 읽기 (단순 / 그룹핑)
  - 값 정규화 (전화번호, 날짜, 면적 등)
"""

import io
import os
import re
import sys
import zipfile
from collections import OrderedDict
from datetime import datetime

# Windows stdout 인코딩 문제 방지
# 라이브러리로 import 시에는 stdout을 건드리지 않음 (호출 스크립트에서 직접 설정)
# 이 모듈을 직접 실행할 때만 적용
if __name__ == "__main__" and sys.platform == "win32" and hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")


# ──────────────────────────────────────────────
#  HWPX (ZIP + XML) 조작
# ──────────────────────────────────────────────

def read_hwpx_xml(hwpx_path, part_name="Contents/section0.xml"):
    """HWPX 파일(ZIP)에서 지정된 XML 파트를 읽어 문자열로 반환.

    Args:
        hwpx_path: HWPX 파일 경로 (str 또는 Path)
        part_name: 읽을 XML 파트 이름 (기본: Contents/section0.xml)

    Returns:
        XML 문자열 (str)
    """
    hwpx_path = str(hwpx_path)
    with zipfile.ZipFile(hwpx_path, "r") as zf:
        return zf.read(part_name).decode("utf-8")


def write_xml_to_hwpx(hwpx_path, part_name, xml_content):
    """수정된 XML을 HWPX 파일(ZIP) 안에 다시 써넣기.

    임시 파일로 작성 후 os.replace()로 원본을 교체하는 안전한 방식.

    Args:
        hwpx_path: HWPX 파일 경로 (str 또는 Path)
        part_name: 대상 XML 파트 이름
        xml_content: 새 XML 내용 (str)
    """
    hwpx_path = str(hwpx_path)
    temp_path = hwpx_path + ".tmp"
    with zipfile.ZipFile(hwpx_path, "r") as zin:
        with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == part_name:
                    zout.writestr(item, xml_content.encode("utf-8"))
                else:
                    zout.writestr(item, zin.read(item.filename))
    os.replace(temp_path, hwpx_path)


# ──────────────────────────────────────────────
#  XML 유틸리티
# ──────────────────────────────────────────────

def xml_escape(value):
    """XML 특수문자 이스케이프.

    Args:
        value: 원본 값 (어떤 타입이든 str로 변환)

    Returns:
        이스케이프된 문자열
    """
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


# ──────────────────────────────────────────────
#  셀 채우기 (핵심 함수)
# ──────────────────────────────────────────────

def fill_cell_by_addr(xml, col, row, value):
    """cellAddr 좌표를 기반으로 HWPX XML의 특정 셀에 값을 채워넣기.

    동작 방식:
    1. `<hp:cellAddr colAddr="col" rowAddr="row"/>`를 찾음
    2. 해당 셀의 `<hp:tc>` 블록을 추출
    3. 빈 self-closing run `<hp:run .../>` 을 찾아서 텍스트를 삽입
    4. 빈 run이 없으면, 기존 `<hp:t>` 텍스트를 교체 (합계 "0" 등)

    Args:
        xml: 전체 section XML 문자열
        col: 열 좌표 (int)
        row: 행 좌표 (int)
        value: 채울 값 (str, 자동으로 XML 이스케이프됨)

    Returns:
        수정된 XML 문자열. 셀을 찾지 못하면 원본 반환.
    """
    target = f'<hp:cellAddr colAddr="{col}" rowAddr="{row}"/>'
    pos = xml.find(target)
    if pos == -1:
        return xml

    tc_start = xml.rfind("<hp:tc", 0, pos)
    tc_end = xml.find("</hp:tc>", pos)
    if tc_start == -1 or tc_end == -1:
        return xml
    tc_end += len("</hp:tc>")

    cell_content = xml[tc_start:tc_end]
    escaped_value = xml_escape(value)

    # 빈 self-closing run을 찾아서 텍스트 삽입
    def replace_once(match):
        run_tag = match.group(0)
        run_open = run_tag[:-2] + ">"  # '/>' → '>'
        return f"{run_open}<hp:t>{escaped_value}</hp:t></hp:run>"

    new_cell = re.sub(r"<hp:run\b[^>]*/>", replace_once, cell_content, count=1)

    if new_cell == cell_content:
        # Fallback: 기존 텍스트가 있는 셀 (예: 합계 "0")
        new_cell = re.sub(
            r"<hp:t>[^<]*</hp:t>",
            f"<hp:t>{escaped_value}</hp:t>",
            cell_content,
            count=1,
        )
        if new_cell == cell_content:
            return xml

    return xml[:tc_start] + new_cell + xml[tc_end:]


# ──────────────────────────────────────────────
#  엑셀 데이터 읽기
# ──────────────────────────────────────────────

def read_excel_data(excel_path, sheet_index=0, sheet_name=None):
    """엑셀 파일을 읽어 (headers, rows) 반환.

    Args:
        excel_path: .xlsx 파일 경로
        sheet_index: 시트 인덱스 (기본 0, sheet_name이 없을 때 사용)
        sheet_name: 시트 이름 (지정하면 sheet_index 무시)

    Returns:
        (headers, rows) 튜플
        - headers: 열 이름 리스트 (1행 기준)
        - rows: 딕셔너리 리스트 [{header: value, ...}, ...]
    """
    from openpyxl import load_workbook

    excel_path = str(excel_path)
    wb = load_workbook(excel_path, data_only=True)

    if sheet_name:
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[sheet_index]

    # 헤더 읽기 (1행)
    headers = []
    for cell in ws[1]:
        val = cell.value
        headers.append(str(val).strip() if val is not None else f"col_{cell.column}")

    # 데이터 행 읽기 (2행부터)
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # 첫 번째 셀이 비어있으면 데이터 끝으로 간주
        if row[0] is None:
            break
        row_dict = {}
        for i, val in enumerate(row):
            if i < len(headers):
                row_dict[headers[i]] = val
        rows.append(row_dict)

    wb.close()
    return headers, rows


def read_excel_grouped(excel_path, group_key, sheet_index=0, sheet_name=None):
    """엑셀 데이터를 특정 키로 그룹핑하여 읽기.

    Args:
        excel_path: .xlsx 파일 경로
        group_key: 그룹핑할 열 이름 (예: '경영체등록번호')
        sheet_index: 시트 인덱스
        sheet_name: 시트 이름

    Returns:
        (headers, groups) 튜플
        - headers: 열 이름 리스트
        - groups: OrderedDict {key_value: [row_dict, ...]}
    """
    headers, rows = read_excel_data(excel_path, sheet_index, sheet_name)

    groups = OrderedDict()
    for row in rows:
        key = row.get(group_key, "")
        if key is None:
            key = ""
        key = str(key).strip()
        if key not in groups:
            groups[key] = []
        groups[key].append(row)

    return headers, groups


# ──────────────────────────────────────────────
#  값 정규화
# ──────────────────────────────────────────────

def normalize_phone(value):
    """전화번호 정규화: 선행 0 복원 + 하이픈 포맷.

    엑셀에서 전화번호가 숫자로 저장되면 앞의 0이 빠지는 문제를 처리.

    Args:
        value: 원본 값 (int, float, str 등)

    Returns:
        정규화된 전화번호 문자열. 빈 값이면 빈 문자열.
    """
    if value is None:
        return ""
    s = str(value).strip()
    if not s or s.lower() == "none":
        return ""

    # Excel의 소수점 (.0) 제거
    if re.fullmatch(r"\d+(\.0+)?", s):
        s = s.split(".")[0]

    digits = re.sub(r"\D", "", s)

    # 선행 0이 빠졌으면 복원
    if digits and not digits.startswith("0") and len(digits) in (9, 10):
        digits = "0" + digits

    # 하이픈 포맷팅
    if len(digits) == 11:
        return f"{digits[:3]}-{digits[3:7]}-{digits[7:]}"
    if len(digits) == 10:
        return f"{digits[:3]}-{digits[3:6]}-{digits[6:]}"
    return s


def normalize_date(value):
    """날짜 값 정규화: datetime → 문자열 변환.

    openpyxl이 datetime 객체를 반환할 수 있으므로 문자열로 변환.

    Args:
        value: 원본 값 (datetime, int, float, str 등)

    Returns:
        문자열. 빈 값이면 빈 문자열.
    """
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    if not s or s.lower() == "none":
        return ""
    # Excel의 소수점 (.0) 제거
    if re.fullmatch(r"\d+(\.0+)?", s):
        s = s.split(".")[0]
    return s


def normalize_area(value):
    """면적 값을 float으로 변환.

    Args:
        value: 원본 값

    Returns:
        float. 변환 실패 시 0.0
    """
    if value is None:
        return 0.0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def area_text(area):
    """면적 float을 사람이 읽기 좋은 문자열로 변환.

    정수면 소수점 없이, 소수면 불필요한 0 제거.

    Args:
        area: 면적 (float)

    Returns:
        포맷된 문자열
    """
    if abs(area - round(area)) < 1e-9:
        return str(int(round(area)))
    return f"{area:.2f}".rstrip("0").rstrip(".")


# ──────────────────────────────────────────────
#  파일명 유틸리티
# ──────────────────────────────────────────────

def sanitize_filename(name):
    r"""Windows에서 사용할 수 없는 파일명 문자 제거.

    \/:*?"<>| 를 _ 로 대체.

    Args:
        name: 원본 파일명

    Returns:
        안전한 파일명 문자열
    """
    return re.sub(r'[\\/:*?"<>|]', "_", str(name))
